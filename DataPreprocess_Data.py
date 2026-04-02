import xlrd
import xlwt
import openpyxl
import math
import linecache
import random
import torch
import csv

import numpy as np
import pandas as pd
import scipy.sparse as sp


lst_weights = []    #存储权重的列表，维度等于窗口大小(4300)
lst = []            #存储当前窗口数值

window_size = 4300     #窗口大小
balance_factor = 4     #e指数分子的平衡系数a

label_index_in = 52    #读取文件标签的列序号
label_index_out = 1    #写入文件标签的列序号

device_indexes = [1,2,3,4,6,7,8,9,10,13,15,17,18,19,20,21,22,23,25,26,27,28,29,35,36,37,39,40,41,42,45,46,47,48,50]

def compute_weights ():

    print ("Computing Weights...")

    #本函数计算窗口内的元素权重
    s = 0
    for i in range (window_size):
        #计算e指数
        val = math.exp (-((window_size - 1 - i) * balance_factor) / window_size)
        lst_weights.append (val)
        s += val

    #print ("Weights before normalization:")
    #for weight in lst_weights:  
    #    print (weight)
        
    #Weight normalization
    for i in range (len (lst_weights)):
        #归一化
        lst_weights[i] = lst_weights[i] / s

    #print ("Weights after normalization:")
    #for weight in lst_weights:
    #    print (weight)


def dot_product ():

    #计算当前窗口数值向量和权重向量的数量积，即窗口内数值加权平均
    result_array = []
    s = 0
    for i in range (len (lst)):
        s += lst[i] * lst_weights[i]
        if (i + 1) % 100 == 0:
            result_array.append (s)
            s = 0
    
    return result_array


def feature_processing ():

    compute_weights ()
    print ("Weights obtained...")
        
    print ("Extracting Features...")
    
    #读取原始数据SWaT_Dataset_Attack_v0.xlsx
    workbook = xlrd.open_workbook (r"SWaT_Dataset_Attack_v0.xlsx", formatting_info = False)

    #print ("All sheets: ", workbook.sheet_names())
    sheet1 = workbook.sheet_by_index (0)

    nrows, ncols = sheet1.nrows, sheet1.ncols
    print ("Sheet Name: %s, #rows: %d, #cols: %d" % (sheet1.name, nrows, ncols))

    #writeexcel = openpyxl.Workbook ()
    #sheet2 = writeexcel.create_sheet (u"Preprocessed_features", 0)

    output_file_index = 0
    #以下循环对原始数据每一列（共51列，实际处理35列）进行滑动窗口加权平均处理，处理全部原始数据    
    for j in device_indexes:
        
        print ("Start processing column %d: %s" % (j, sheet1.cell (1, j).value))

        #首先获取当前列的初始窗口，即包含前window_size个数值的列表
        #Initialize the 1st window of a new column
        lst.clear ()
        for k in range (2, window_size + 2):
            lst.append (sheet1.cell (k, j).value)

        #滑动窗口，计算窗口内数值加权平均值，写入输出文档相应位置
        #Window Sliding, i points to the last element in the current window
        print ("\tInitial window ready...")

        fname = 'Preprocessed_Features_Device_' + str (output_file_index) + ".csv"
        f = open (fname, 'w')
        
        for i in range (window_size + 1, nrows):
            result = dot_product ()
            #Write result to new file

            rowstr = ""
            for k in range (len (result)):
                rowstr = rowstr + str (result [k])
                if k != len (result) - 1:
                    rowstr = rowstr + ","
                else:
                    rowstr = rowstr + "\n"
            
            f.write (rowstr)
            
            #sheet2.cell (i + 1 - window_size, j + 1, result)
            if i != nrows - 1:
                lst.pop (0)
                lst.append (sheet1.cell (i + 1, j).value)

        f.close ()
        output_file_index += 1
        print ("\tProcessed Features for Current Device Saved...")
        #if output_file_index >= 10:
        #    print ("Breaking out of loop...")
        #    break

    #保存写入的文档
    #print ("Start Saving...")
    #writeexcel.save ("Preprocessed_Features.xlsx")
    print ("Processed Features Saved to New File")

'''
Modified: Producing .csv File instead of .xlsx 2022/9/27
'''
def label_processing ():

    #滑动窗口获取标签，同一窗口只要出现Attack就标为1（Abnormal）,否则为0（Normal），处理全部原始数据
    #0 - normal, 1 - abnormal
    print ("Extracting Labels...")
    workbook = xlrd.open_workbook (r"SWaT_Dataset_Attack_v0.xlsx", formatting_info = False)

    #print ("All sheets: ", workbook.sheet_names())
    sheet1 = workbook.sheet_by_index (0)

    nrows, ncols = sheet1.nrows, sheet1.ncols
    print ("Sheet1 Name: %s, #rows: %d, #cols: %d" % (sheet1.name, nrows, ncols))

    #writeexcel = openpyxl.Workbook ()
    #sheet2 = writeexcel.create_sheet (u"Labeled_data", 0)

    fname = 'Preprocessed_Labels.csv'
    f = open (fname, 'w')

    nnormals, nattacks = 0, 0
    
    #find the number of Normal / Attack Labels in the first window
    #获取初始滑动窗口内0/1标签的数量
    
    for k in range (2, window_size + 2):

        if (sheet1.cell (k, label_index_in).value [0] == 'N'):
            nnormals += 1
        else:
            nattacks += 1

    print ("Stats in initial window: Normal states (%d), Attack states (%d)" % (nnormals, nattacks))

    #i points to the last element in the current window
    #滑动窗口，为第i个窗口标注标签，
    for i in range (window_size + 1, nrows):

        #判断当前窗口标签并写入到输出文件里
        if nattacks > 0:
            label = 1
        else:
            label = 0
        row = str (label) + "\n"
        f.write (row)
            
        if i != nrows - 1:
            #获取当前窗口即将出队和入队的标签，通过这两个标签更新下一个窗口内0/1标签数量
            front, rear = sheet1.cell (i - window_size + 1, label_index_in).value, sheet1.cell (i + 1, label_index_in).value

            if front [0] == "N":
                nnormals -= 1
            else:
                nattacks -= 1

            if rear [0] == "N":
                nnormals += 1
            else:
                nattacks += 1
    
    #print ("Start Saving...")
    #writeexcel.save ("Preprocessed_Labels.xlsx")
    print ("Closing File...")
    f.close ()
    print ("Preprocessed Labels Saved to New File")

def down_sampling ():

    #List of files whose data is sampled. Normalization is completed as well.
    files_src, files_dst = [], []
    for i in range (len (device_indexes)):
        files_src.append ('Preprocessed_Features_Device_' + str (i) + ".csv")
        files_dst.append ('Preprocessed_Downsampled_Features_Device_' + str (i) + ".csv")

    files_src.append ('Preprocessed_Labels.csv')
    files_dst.append ('Preprocessed_Downsampled_Labels.csv')

    print (len (files_src), len (files_dst))
    print (files_src)
    print (files_dst)

    s = 0
    for i in range (window_size):
        val = math.exp (-((window_size - 1 - i) * balance_factor) / window_size)
        s += val

    s = s * 100 / window_size
    print ("Sum: ", s)

    for i in range (len (files_src)):
        counter = 0
        f = open (files_dst [i], "w")
        with open (files_src [i], "rt") as infile:
            reader = csv.reader (infile, delimiter = ",")
            for row in reader:
                #print (row)
                counter += 1
                if (counter % 10 != 0):
                    continue
            
                output_row = ""
                for j in range (len (row)):
                    #print (float (e))
                    if (i != len (files_src) - 1):
                        new_val = math.atan (float (row [j]) * s)
                    else:
                        new_val = int (row [j])
                
                    #print (new_val)
                    output_row = output_row + str (new_val) + ","

                output_row = output_row [: -1] + "\n"
                #print (output_row)
                f.write (output_row)
        f.close ()
        print (files_dst [i], " created!")


batch_size = 64

def sample_shuffle ():

    #nnodes_3_layer = 35 + 6 + 1
    #nnodes_2_layer = 35 + 6

    nlines = 44544    #Number of lines in each file to be aggregated, denoting the number of time ticks taken into account
    ndevices = 35

    random_lst = []

    features_fnames = []
    label_fname = 'Preprocessed_Downsampled_Labels.csv'

    frows = []
    flabel_handle = open (label_fname, 'rt')

    for i in range (ndevices):
        print ("Producing List for Device ", i)
        name = "Preprocessed_Downsampled_Features_Device_" + str (i) + ".csv"
        features_fnames.append (name)
        f_curr = open (name, 'rt')
        rd = csv.reader (f_curr)
        rows = list (rd)
        frows.append (rows)
    #print (features_fnames)

    f_labels = open (label_fname, 'rt')
    rd_labels = csv.reader (f_labels)
    label_rows = list (rd_labels)

    print ("Label list ready!")

    for i in range (nlines):
        random_lst.append (i)
    
    random.shuffle (random_lst)
    print (random_lst)

    f = open ("swat_nodes_3_all_time_ticks_dev_alt.csv", 'w')

    zeroes = ""
    for i in range (43 + 1):
        zeroes = zeroes + "0"
        if i != 42 + 1:
            zeroes = zeroes + ","
        else:
            zeroes = zeroes + "\n"

    print ("Zeroes Vector Initialized...")

    nid = 1
    for i in random_lst:
        #print ("Fetching time tick ", i)
        for k in range (1 + 6):   # 1 CRP + 6 Controllers
            row = str (nid) + "," + zeroes
            f.write (row)
            #print (row)
            nid = nid + 1
        for rows in frows:
            row = rows [i]
            output_row = str (nid) + ","
            
            for j in range (len (row)):
                output_row = output_row + row [j] + ","

            output_row = output_row + str (label_rows [i][0]) + "\n"
            #print (output_row)
            f.write (output_row)
            nid = nid + 1

    f.close ()

from sklearn.decomposition import PCA

def PCA_com ():

    data_node = np.loadtxt('swat_nodes_3_all_time_ticks_dev_alt.csv',delimiter=',')
    print ("Shape loaded: ", data_node.shape)
    X = data_node [:, 1 : -1]
    y = data_node [:, -1]
    print ("X.shape: ", X.shape)
    print ("y.shape: ", y.shape)
    #data_label = np.loadtxt('labels.csv',int,delimiter=',')
    #data_label = np.triu(data_label,1)
    print (type (data_node))
    
    pca = PCA (n_components = 0.99)#表示保留了99%的信息
    low = pca.fit_transform (X).shape [1]#降低后的维数
    Data_pca = PCA (n_components = low)
    X_pca = Data_pca.fit_transform (X)
    print ("Downsized dimension: ", low)

    idx = np.reshape (np.array (range (X_pca.shape [0])), (X_pca.shape [0], 1))
    y = np.reshape (y, (X_pca.shape [0], 1))
    print (idx.shape, X_pca.shape, y.shape)

    PCA_data = np.concatenate ((idx, X_pca, y), axis = 1)
    print ("Aggregated PCA Data shape: ", PCA_data.shape)

    np.savetxt ("swat_nodes_3_pca_all_time_ticks_dev_alt.csv", PCA_data, delimiter = ',')
    #print (PCA_data [0][1],PCA_data [0][2],PCA_data [0][3],PCA_data [0][4],PCA_data [0][5],PCA_data [0][6],PCA_data [0][7])
    print ("PCA Data Saved to New File!")

def count_labels_original ():

    #滑动窗口获取标签，同一窗口只要出现Attack就标为1（Abnormal）,否则为0（Normal），处理全部原始数据
    #0 - normal, 1 - abnormal
    print ("Counting labels in original dataset...")
    workbook = xlrd.open_workbook (r"SWaT_Dataset_Attack_v0.xlsx", formatting_info = False)

    #print ("All sheets: ", workbook.sheet_names())
    sheet1 = workbook.sheet_by_index (0)

    nrows, ncols = sheet1.nrows, sheet1.ncols
    print ("Sheet1 Name: %s, #rows: %d, #cols: %d" % (sheet1.name, nrows, ncols))

    #writeexcel = openpyxl.Workbook ()
    #sheet2 = writeexcel.create_sheet (u"Labeled_data", 0)

    nnormals, nattacks = 0, 0
    
    #find the number of Normal / Attack Labels in the first window
    #获取初始滑动窗口内0/1标签的数量
    
    for k in range (2, nrows):

        if (sheet1.cell (k, label_index_in).value [0] == 'N'):
            nnormals += 1
        else:
            nattacks += 1

    print ("Normal states (%d), Attack states (%d)" % (nnormals, nattacks))
    anomaly_ratio = float (nattacks) / float (nattacks + nnormals)
    print ("Anomaly Ratio: ", anomaly_ratio)

    
def count_labels ():

    print ("Counting labels in preprocessed dataset...")
    f = open ("swat_nodes_3_all_time_ticks_dev_alt.csv", 'rt')
    rd = csv.reader (f)
    rows = list (rd)

    m = 0
    n = 0

    for row in rows:
        if (row [len (row) - 1] == "0"):
            m += 1
        else:
            n += 1

    print ("Number of Zeroes: ", m, "\tNumber of Ones: ", n)
    anomaly_ratio = float (n) / float (m + n)
    print ("Anomaly Ratio: ", anomaly_ratio)
    
#feature_processing ()
#label_processing ()
#down_sampling ()
sample_shuffle ()
PCA_com ()
#count_labels_original ()
#count_labels ()
